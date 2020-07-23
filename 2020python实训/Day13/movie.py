from bs4 import BeautifulSoup
import requests
import openpyxl

def HTMLDownloader(page):
    url = 'https://movie.douban.com/top250?start='+ str(page*25) +'&filter='
    #print(url)
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) \
    AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.129 Safari/537.36'}
    try:
        reponse = requests.get(url,headers = headers)
        if reponse.status_code == 200:#响应成功
            return reponse.text
    except requests.RequestException as e:
        return None

def HTMLParser(reponse):
    """
	网页解析出的电影数据都保存在列表中
	列表的每个元素都包含一部电影的信息
	"""
    outcome = []
    if reponse == None:
        pass
    else:
        soup = BeautifulSoup(reponse,'lxml') #BeautifulSoup构造方法
        #寻找所有电影的div标签
        items = soup.findAll('div',attrs={'class':'item'})
         #从div标签中获取电影的各种信息
        for item in items:
            title = item.find('span',attrs = {'class':'title'}).string
            rank = item.find('span',attrs = {'class':'rating_num'}).string
            movinfo = item.find('p',attrs = {'class':''}).text
            #有的电影没有评论
            if item.find('span',attrs = {'class':'inq'}):
                comment = item.find('span',attrs = {'class':'inq'}).string
            else:
                comment = ''
            print('clawing',title,rank,movinfo,comment)
            outcome.append([title,rank,movinfo.strip(),comment.strip()])
        return outcome

def MovieInfoSaver(mov):
    """
    保存爬虫结果到excel中
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'douban_movie'
    sheet.cell(1,1,'电影名')
    sheet.cell(1,2,'豆瓣评分')
    sheet.cell(1,3,'简介')
    sheet.cell(1,4,'评论')
    for idx,m in enumerate(mov):
        sheet.cell(idx + 2,1,m[0])
        sheet.cell(idx + 2,2,m[1])
        sheet.cell(idx + 2,3,m[2])
        sheet.cell(idx + 2,4,m[3])

    workbook.save(u'douban_hrank_movie.xlsx')  # 保存工作簿

if __name__ == "__main__":
    movinfo = []
    #通过page变量请求10个页面
    for page in range(10):
        reponse = HTMLDownloader(page)
        movinfo += HTMLParser(reponse)
    MovieInfoSaver(movinfo)
